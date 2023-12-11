import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from scipy.stats import norm
from itertools import product

# step 1 -> construct 2 arrays. array1 would be a 2d array with cell references and array 2 would be a 2d array with
# the corresponding sheet references
# probabilities would be given as input
# step 2 -> Using probabilities for columns, sample the columns and rows using uniform distribution. output the indices.
# step 3 -> Using the indices, get the cell references and sheet references as 2 arrays
# step 4 -> Using the cell references and sheet references, highlight the cells in the excel sheet


def create_sheet_cell_references(file_path):
    workbook = openpyxl.load_workbook(file_path)
    cell_references = []
    sheet_references = []
    for sheet in workbook.worksheets:
        columns = [chr(col + 65) for col in range(sheet.max_column)]
        rows = range(2, sheet.max_row + 1)
        cell_references.extend([f"{col}{row}" for row, col in product(rows, columns)])
        sheet_references.extend([sheet.title] * sheet.max_row * sheet.max_column)
    cell_references = np.array(cell_references).reshape(-1, sheet.max_column)
    sheet_references = np.array(sheet_references).reshape(-1, sheet.max_column)
    # cell_references.reshape(-1, sheet.max_column)
    # sheet_references.reshape(-1, sheet.max_column)
    return cell_references, sheet_references, workbook


def sample_cells_with_probabilities(
    cell_references, sheet_references, n, column_probabilities=None, seed=None
):
    if column_probabilities is None:
        column_probabilities = (
            np.ones(cell_references.shape[1]) / cell_references.shape[1]
        )

    # Normalize probabilities in each column
    normalized_probabilities = column_probabilities / column_probabilities.sum()

    # Generate random indices based on column-wise probabilities
    column_indices = np.random.choice(
        cell_references.shape[1], size=n, p=normalized_probabilities
    )

    # Generate random indices for rows
    row_indices = np.random.choice(cell_references.shape[0], size=n)

    # Combine row and column indices to get the final sampled coordinates
    sampled_coordinates = (row_indices, column_indices)

    # Get the sampled elements
    sampled_cell_references = cell_references[sampled_coordinates]
    sampled_sheet_references = sheet_references[sampled_coordinates]

    return sampled_coordinates, sampled_cell_references, sampled_sheet_references


def highlight_cells(worksheet, cells):
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    for cell in cells:
        worksheet[cell].fill = yellow_fill


def highlight_cells_workbook(
    workbook, sampled_cell_references, sampled_sheet_references
):
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for cell in sampled_cell_references[sampled_sheet_references == sheet_name]:
            worksheet[cell].fill = yellow_fill


def random_cells(num_cells, num_rows, num_cols):
    columns = [chr(col + 65) for col in range(num_cols)]  # 'A' to 'Z'
    rows = range(2, num_rows + 1)  # 1 to 25
    cell_references = [f"{col}{row}" for col, row in product(columns, rows)]
    return np.random.choice(cell_references, num_cells, replace=True)


def calculate_sample_size(o, t, confidence):
    z_score = norm.ppf(1 - (1 - confidence) / 2)
    return int(np.ceil((z_score**2 * o * (1 - o)) / ((t - o) ** 2)))


if __name__ == "__main__":
    file_path = "/Users/bhanuteja/Downloads/qc_nov16.xlsx"
    cell_references, sheet_references, workbook = create_sheet_cell_references(
        file_path
    )
    (
        sampled_coordinates,
        sampled_cell_references,
        sampled_sheet_references,
    ) = sample_cells_with_probabilities(cell_references, sheet_references, 777)
    highlight_cells_workbook(
        workbook, sampled_cell_references, sampled_sheet_references
    )
    workbook.save(file_path)


# if __name__ == "__main__":
#     input_file = "/Users/bhanuteja/Downloads/qc_report.xlsx"
#     output_file = "/Users/bhanuteja/Downloads/qc_report.xlsx"
#     observed_error_rate = 0.01  # Replace with your observed error rate
#     target_error_rate = 0.003  # Replace with your target error rate
#     confidence = 0.95  # Replace with your desired confidence interval
#     print(f"Sample size is {calculate_sample_size(observed_error_rate, target_error_rate, confidence)}")
#     # print(f"{random_cells(10, 100, 10)}")
#     # main(input_file, output_file, observed_error_rate, target_error_rate, confidence)
